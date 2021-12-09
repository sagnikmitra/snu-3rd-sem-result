import streamlit as st
import pandas as pd
import openpyxl as pxl
import base64
cse = pd.read_excel('cse.xlsx')
exp = pxl.load_workbook('cse.xlsx')
sheet = exp['CSE']
branch = st.selectbox('Select your Branch', cse['Dept'].unique())
roll = st.text_input('Enter your Enrollment Number')
reg = st.text_input('Enter your Registration Number')


def show_pdf(file_path):
    with open(file_path, "rb") as f:
        base64_pdf = base64.b64encode(f.read()).decode('utf-8')
    pdf_display = '<iframe src="data:application/pdf;base64,{base64_pdf}" width="700" height="1000" type="application/pdf"></iframe>'
    # pdf_display = f'<embed src="data:application/pdf;base64,{base64_pdf}" width="700" height="1000" type="application/pdf">'
    st.markdown(pdf_display, unsafe_allow_html=True)


for i in range(2, sheet.max_row+1):
    if sheet.cell(row=i, column=5).value.lower() == roll.lower():
        if sheet.cell(row=i, column=6).value.lower() == reg.lower():
            if sheet.cell(row=i, column=2).value.lower() == branch.lower():
                strPdf = './Results/' + \
                    sheet.cell(row=i, column=5).value.lower()+'-2.pdf'
                show_pdf(strPdf)
